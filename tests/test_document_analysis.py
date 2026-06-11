from io import BytesIO

from openpyxl import Workbook

from ojtflow.data_tools.document_analysis import (
    analyze_pdf,
    analyze_workbook,
    build_document_intelligence_profile,
)


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Labs"
    sheet.append(["patient_id", "lab_name", "value", "unit"])
    sheet.append(["P001", "HbA1c", 7.4, "%"])
    sheet.merge_cells("B4:C4")
    hidden = workbook.create_sheet("Hidden Notes")
    hidden.sheet_state = "hidden"
    hidden.append(["note", "owner"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_workbook_profile_detects_headers_merged_cells_and_hidden_sheets() -> None:
    profile = analyze_workbook(
        data=_workbook_bytes(),
        filename="labs.xlsx",
        source_format="xlsx",
    )

    assert profile is not None
    assert profile.sheet_count == 2
    assert profile.hidden_sheet_count == 1
    assert profile.sheets[0].header_row == 1
    assert profile.sheets[0].headers[:4] == ["patient_id", "lab_name", "value", "unit"]
    assert profile.sheets[0].merged_cell_ranges == ["B4:C4"]
    assert any("hidden" in warning.lower() for warning in profile.warnings)


def test_pdf_profile_returns_clear_warning_when_analyzer_unavailable_or_fails() -> None:
    profile = analyze_pdf(data=b"not a real pdf", filename="scan.pdf")

    assert profile.analyzer in {"unavailable", "pymupdf", "pypdf"}
    assert profile.scan_likelihood == "unknown"
    assert profile.warnings


def test_document_intelligence_profile_scores_empty_text_as_poor() -> None:
    profile = build_document_intelligence_profile(
        data=b"fake",
        filename="scan.pdf",
        source_format="pdf",
        extracted_text="",
        extractor_used="markitdown",
        extraction_confidence=0.2,
        extraction_warnings=["markitdown returned empty text."],
    )

    assert profile.quality.level == "poor"
    assert profile.quality.requires_review is True
    assert "empty_text" in profile.quality.factors
    assert profile.explanation.needs_review
    assert profile.explanation.limitations


def test_document_quality_flags_fallback_and_ambiguous_tables() -> None:
    profile = build_document_intelligence_profile(
        data=_workbook_bytes(),
        filename="labs.xlsx",
        source_format="xlsx",
        extracted_text="patient_id,lab_name,value,unit\nP001,HbA1c,7.4,%\n",
        extractor_used="markitdown",
        extraction_confidence=0.92,
        extraction_warnings=["markitdown failed once before fallback succeeded."],
    )

    assert "extractor_conflict_or_fallback" in profile.quality.factors
    assert "malformed_or_ambiguous_tables" in profile.quality.factors
    assert profile.quality.requires_review is True
