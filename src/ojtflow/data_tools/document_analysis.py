"""Document intelligence helpers for spreadsheet/PDF extraction quality."""

from __future__ import annotations

import io
from typing import Any

from ojtflow.core.contracts.artifacts import (
    DocumentIntelligenceProfile,
    ExtractionExplanation,
    ExtractionQualityReport,
    PdfContentProfile,
    WorkbookParsingProfile,
    WorkbookSheetProfile,
)


WORKBOOK_FORMATS = {"xlsx", "xls"}


def build_document_intelligence_profile(
    *,
    data: bytes,
    filename: str,
    source_format: str,
    extracted_text: str,
    extractor_used: str,
    extraction_confidence: float,
    extraction_warnings: list[str],
) -> DocumentIntelligenceProfile:
    """Build workbook/PDF profile, quality score, and user-facing explanation."""

    workbook = analyze_workbook(data=data, filename=filename, source_format=source_format)
    pdf = analyze_pdf(data=data, filename=filename) if source_format == "pdf" else None
    quality = build_extraction_quality(
        extracted_text=extracted_text,
        extraction_confidence=extraction_confidence,
        extraction_warnings=extraction_warnings,
        workbook=workbook,
        pdf=pdf,
    )
    explanation = build_extraction_explanation(
        filename=filename,
        source_format=source_format,
        extractor_used=extractor_used,
        extracted_text=extracted_text,
        workbook=workbook,
        pdf=pdf,
        quality=quality,
    )
    return DocumentIntelligenceProfile(
        source_format=source_format,
        workbook=workbook,
        pdf=pdf,
        quality=quality,
        explanation=explanation,
    )


def analyze_workbook(
    *,
    data: bytes,
    filename: str,
    source_format: str,
) -> WorkbookParsingProfile | None:
    """Profile spreadsheet workbooks without changing extracted text."""

    if source_format not in WORKBOOK_FORMATS:
        return None
    if source_format == "xls":
        return WorkbookParsingProfile(
            filename=filename,
            source_format=source_format,
            warnings=[
                "Legacy .xls workbook profiling is not enabled; convert to .xlsx "
                "or add an xlrd-backed analyzer."
            ],
        )

    try:
        from openpyxl import load_workbook  # type: ignore[import]
    except ImportError:
        return WorkbookParsingProfile(
            filename=filename,
            source_format=source_format,
            warnings=["openpyxl is not installed; workbook sheet profiling was skipped."],
        )

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=False, data_only=True)
    except Exception as exc:
        return WorkbookParsingProfile(
            filename=filename,
            source_format=source_format,
            warnings=[f"Workbook profiling failed: {type(exc).__name__}."],
        )

    sheets: list[WorkbookSheetProfile] = []
    warnings: list[str] = []
    for index, worksheet in enumerate(workbook.worksheets):
        sheet_warnings: list[str] = []
        hidden_state = str(getattr(worksheet, "sheet_state", "visible") or "visible")
        if hidden_state != "visible":
            sheet_warnings.append(f"Sheet '{worksheet.title}' is {hidden_state}.")
        merged_ranges = [str(range_ref) for range_ref in worksheet.merged_cells.ranges]
        if merged_ranges:
            sheet_warnings.append(
                f"Sheet '{worksheet.title}' contains {len(merged_ranges)} merged cell range(s)."
            )
        header_row, headers = _detect_header_row(worksheet)
        if header_row is None and worksheet.max_row:
            sheet_warnings.append(f"Sheet '{worksheet.title}' has no clear header row.")
        warnings.extend(sheet_warnings)
        sheets.append(
            WorkbookSheetProfile(
                sheet_name=str(worksheet.title),
                sheet_index=index,
                row_count=int(worksheet.max_row or 0),
                column_count=int(worksheet.max_column or 0),
                hidden_state=hidden_state,
                header_row=header_row,
                headers=headers,
                merged_cell_ranges=merged_ranges,
                warnings=sheet_warnings,
            )
        )

    hidden_count = sum(1 for sheet in sheets if sheet.hidden_state != "visible")
    return WorkbookParsingProfile(
        filename=filename,
        source_format=source_format,
        sheet_count=len(sheets),
        visible_sheet_count=len(sheets) - hidden_count,
        hidden_sheet_count=hidden_count,
        sheets=sheets,
        warnings=warnings,
    )


def analyze_pdf(*, data: bytes, filename: str) -> PdfContentProfile:
    """Detect digital/scanned PDF characteristics using optional local parsers."""

    fitz_profile = _analyze_pdf_with_fitz(data=data, filename=filename)
    if fitz_profile is not None:
        return fitz_profile
    pypdf_profile = _analyze_pdf_with_pypdf(data=data, filename=filename)
    if pypdf_profile is not None:
        return pypdf_profile
    return PdfContentProfile(
        filename=filename,
        scan_likelihood="unknown",
        requires_ocr=False,
        analyzer="unavailable",
        warnings=[
            "PDF scanned-vs-digital detection skipped because PyMuPDF or pypdf "
            "is not installed."
        ],
    )


def build_extraction_quality(
    *,
    extracted_text: str,
    extraction_confidence: float,
    extraction_warnings: list[str],
    workbook: WorkbookParsingProfile | None,
    pdf: PdfContentProfile | None,
) -> ExtractionQualityReport:
    """Score extraction quality from deterministic signals."""

    score = 1.0
    factors: list[str] = []
    warnings: list[str] = []
    if not extracted_text.strip():
        score -= 0.75
        factors.append("empty_text")
        warnings.append("Extractor produced empty text.")
    elif len(extracted_text.strip()) < 40:
        score -= 0.2
        factors.append("very_short_text")
    if extraction_confidence < 0.75:
        score -= 0.2
        factors.append("low_extractor_confidence")
    if extraction_warnings:
        score -= min(0.25, len(extraction_warnings) * 0.05)
        factors.append("extractor_warnings")
        warnings.extend(extraction_warnings)
    if any(_looks_like_extractor_conflict(warning) for warning in extraction_warnings):
        score -= 0.1
        factors.append("extractor_conflict_or_fallback")
    if pdf and pdf.requires_ocr:
        score -= 0.25
        factors.append("pdf_requires_ocr")
        warnings.extend(pdf.warnings)
    if workbook and workbook.warnings:
        score -= min(0.25, len(workbook.warnings) * 0.05)
        factors.append("workbook_structure_warnings")
        if any(_looks_like_table_structure_problem(warning) for warning in workbook.warnings):
            factors.append("malformed_or_ambiguous_tables")
        warnings.extend(workbook.warnings)
    score = round(max(0.0, min(1.0, score)), 3)
    if score >= 0.8:
        level = "good"
    elif score >= 0.45:
        level = "review"
    else:
        level = "poor"
    return ExtractionQualityReport(
        score=score,
        level=level,
        requires_review=level != "good" or bool(warnings),
        factors=sorted(set(factors)),
        warnings=_dedupe(warnings),
    )


def build_extraction_explanation(
    *,
    filename: str,
    source_format: str,
    extractor_used: str,
    extracted_text: str,
    workbook: WorkbookParsingProfile | None,
    pdf: PdfContentProfile | None,
    quality: ExtractionQualityReport,
) -> ExtractionExplanation:
    """Create operator-facing extraction explanation fragments."""

    read = [
        f"Read {len(extracted_text)} character(s) from {filename} using {extractor_used}."
    ]
    skipped: list[str] = []
    needs_review: list[str] = []
    limitations: list[str] = []

    if workbook:
        read.append(
            f"Profiled {workbook.sheet_count} workbook sheet(s); "
            f"{workbook.hidden_sheet_count} hidden."
        )
        if workbook.hidden_sheet_count:
            needs_review.append("Hidden workbook sheets may contain data not expected by users.")
        if any(sheet.merged_cell_ranges for sheet in workbook.sheets):
            needs_review.append("Merged spreadsheet cells can shift headers or table values.")
        if source_format == "xls":
            skipped.append("Detailed legacy .xls sheet profiling was skipped.")
    else:
        skipped.append("Workbook profiling did not apply to this file type.")

    if pdf:
        read.append(
            f"Profiled {pdf.page_count} PDF page(s); scan likelihood is {pdf.scan_likelihood}."
        )
        if pdf.requires_ocr:
            needs_review.append("PDF appears to require OCR for complete extraction.")
        if pdf.analyzer == "unavailable":
            skipped.append("PDF scanned-vs-digital detection was skipped.")
    elif source_format != "pdf":
        skipped.append("PDF scanned-vs-digital detection did not apply to this file type.")

    if quality.requires_review:
        needs_review.append(f"Extraction quality is {quality.level} ({quality.score}).")
    if not extracted_text.strip():
        limitations.append("No text was extracted; downstream validation may miss all source data.")
    limitations.extend(quality.warnings)
    return ExtractionExplanation(
        read=_dedupe(read),
        skipped=_dedupe(skipped),
        needs_review=_dedupe(needs_review),
        limitations=_dedupe(limitations),
    )


def _detect_header_row(worksheet) -> tuple[int | None, list[str]]:
    best_row: int | None = None
    best_values: list[str] = []
    best_score = 0
    max_rows = min(int(worksheet.max_row or 0), 10)
    for row_index in range(1, max_rows + 1):
        values = [
            _cell_text(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, int(worksheet.max_column or 0) + 1)
        ]
        non_empty = [value for value in values if value]
        unique = len(set(value.casefold() for value in non_empty))
        score = unique + sum(1 for value in non_empty if not _looks_numeric(value))
        if unique >= 2 and score > best_score:
            best_score = score
            best_row = row_index
            best_values = non_empty
    return best_row, best_values


def _analyze_pdf_with_fitz(*, data: bytes, filename: str) -> PdfContentProfile | None:
    try:
        import fitz  # type: ignore[import]
    except ImportError:
        return None
    try:
        doc = fitz.open(stream=data, filetype="pdf")
        page_count = len(doc)
        digital_text_pages = 0
        image_pages = 0
        image_only_pages = 0
        for page in doc:
            text = (page.get_text("text") or "").strip()
            images = page.get_images(full=True)
            if len(text) >= 20:
                digital_text_pages += 1
            if images:
                image_pages += 1
            if images and len(text) < 20:
                image_only_pages += 1
        doc.close()
    except Exception as exc:
        return PdfContentProfile(
            filename=filename,
            analyzer="pymupdf",
            scan_likelihood="unknown",
            warnings=[f"PyMuPDF PDF analysis failed: {type(exc).__name__}."],
        )
    return _pdf_profile(
        filename=filename,
        analyzer="pymupdf",
        page_count=page_count,
        digital_text_pages=digital_text_pages,
        image_pages=image_pages,
        image_only_pages=image_only_pages,
    )


def _analyze_pdf_with_pypdf(*, data: bytes, filename: str) -> PdfContentProfile | None:
    try:
        from pypdf import PdfReader  # type: ignore[import]
    except ImportError:
        return None
    try:
        reader = PdfReader(io.BytesIO(data))
        page_count = len(reader.pages)
        digital_text_pages = sum(
            1 for page in reader.pages if len((page.extract_text() or "").strip()) >= 20
        )
    except Exception as exc:
        return PdfContentProfile(
            filename=filename,
            analyzer="pypdf",
            scan_likelihood="unknown",
            warnings=[f"pypdf PDF analysis failed: {type(exc).__name__}."],
        )
    return _pdf_profile(
        filename=filename,
        analyzer="pypdf",
        page_count=page_count,
        digital_text_pages=digital_text_pages,
        image_pages=0,
        image_only_pages=max(0, page_count - digital_text_pages),
        image_count_unknown=True,
    )


def _pdf_profile(
    *,
    filename: str,
    analyzer: str,
    page_count: int,
    digital_text_pages: int,
    image_pages: int,
    image_only_pages: int,
    image_count_unknown: bool = False,
) -> PdfContentProfile:
    warnings: list[str] = []
    if page_count == 0:
        scan_likelihood = "unknown"
        requires_ocr = False
        warnings.append("PDF has no pages.")
    elif digital_text_pages == page_count:
        scan_likelihood = "digital"
        requires_ocr = False
    elif digital_text_pages == 0 and (image_pages > 0 or image_count_unknown):
        scan_likelihood = "scanned"
        requires_ocr = True
        warnings.append("PDF appears scanned or image-only; OCR is required.")
    elif image_only_pages > 0:
        scan_likelihood = "mixed"
        requires_ocr = True
        warnings.append("PDF has a mix of digital text and pages that may need OCR.")
    else:
        scan_likelihood = "unknown"
        requires_ocr = False
    if image_count_unknown:
        warnings.append("PDF image-page counts are approximate because pypdf cannot inspect images.")
    return PdfContentProfile(
        filename=filename,
        page_count=page_count,
        digital_text_pages=digital_text_pages,
        image_pages=image_pages,
        image_only_pages=image_only_pages,
        scan_likelihood=scan_likelihood,
        requires_ocr=requires_ocr,
        analyzer=analyzer,
        warnings=warnings,
    )


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_numeric(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False


def _looks_like_extractor_conflict(warning: str) -> bool:
    normalized = warning.casefold()
    return any(term in normalized for term in ("fallback", "falling back", "failed"))


def _looks_like_table_structure_problem(warning: str) -> bool:
    normalized = warning.casefold()
    return any(term in normalized for term in ("merged", "header", "malformed", "hidden"))


def _dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result
